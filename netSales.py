#bu çalışmayla birlikte ilk kez diyez kullanımına başlamış bulunuyorum.

import openpyxl #her şeyin başı bu. bununla birlikte excel komutlarını kullanmaya açmış oluyorum, o kütüphaneye giriyorum yani aslında.

#kodları videodan bağımsız bir şekilde kendim yazmaya çalışıyorum. video linki: https://www.youtube.com/watch?v=hFDrWvDOYFA

Workbook = openpyxl.load_workbook("C:\\Users\\kadir\\Desktop\\netSales.xlsx")  
Worksheet = Workbook.active

 #belirttiğim satır ve sütundaki değerleri yazmasını istedim.

max_column = Worksheet.max_column
max_row = Worksheet.max_row             # maksimum satır ve sütun sayısını tanımlayarak sayılarını yazmasını istedim.

print("total rows:",max_row)
print("total columns", max_column)

print("Values of the first column")
for i in range(1, 7):                               #1. sütundaki 1'den 6'ya kadarki olan değerleri yazdırdım.
    cell = Worksheet.cell (row=i, column=1)
    print(cell.value)

print("Values of the first row")
for i in range(1, 5):
    cell = Worksheet.cell (row=2, column=i)
    print(cell.value)                           #2.satırdaki 1'den 4'e kadarki olan değerleri yazdırdım.
    
     