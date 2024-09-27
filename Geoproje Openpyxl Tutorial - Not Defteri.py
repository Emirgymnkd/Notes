#BURADAKİ KODLAR ELBETTE BİR BÜTÜN HALİNDE ÇALIŞMAYACAK AMA GEREKLİ KOMBİNASYONLAR YAPILDIĞINDA ÇALIŞACAKTIR. HİÇBİR KODU DENEMEDEN VEYA HATALI BİR ŞEKİLDE KOYMUYORUM. BURASI SADECE BİR LİSTE.

"""import openpyxl"""

"""Workbook = openpyxl.load_workbook("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Openpyxl Tutorial\\GeoProje Openpyxl Tutorial.xlsx")"""


"""print(Workbook.sheetnames)"""  #sayfa ismini görüntüler. 

# spesifik bir sayfa ismini görüntülemek istiyorsam:
"""worksheet = Workbook["Score"]
print(worksheet)"""

"""worksheet = Workbook["Sheet1"]
print(worksheet)"""

#yeni bir sayfa oluşturmak istiyorum ve ismi de new sheet olsun:
"""Workbook.create_sheet("New Sheet")""" 

#oluşturacağım sayfanın konumunu ayarlamak istiyorum:
"""Workbook.create_sheet("New Sheet", index:1)"""   #-----> 1. sayfadan sonra koy demek oluyor, en başa gelmesini istiyorsan 0 yazmalısın. index koymazsan direkt en sona koyar.

#spesifik bir hücreyi istiyorum:
"""worksheet = Workbook["Sheet1"]
print(worksheet["B1"])""" #<Cell 'Sheet1'.B1> sonucuyla karşılaştım. 

#spesifik bir hücrenin değerini istiyorum:
"""worksheet = Workbook["Sheet1"]
print(worksheet["B1"].value)""" #Balance sonucuyla karşılaştım.

#spesifik bir hücrenin değerini değiştirmek istiyorum:
"""worksheet = Workbook["Sheet1"]
worksheet['B5'].value""" = 8

#veya:
"""worksheet['A4']""" = 'kadiremir'

#veya:
"""print(worksheet.cell(row=6, column=1).value)""" = 8 #yazabilirim

#belirli bir sütundaki değerleri liste halinde yazdırmak istiyorum.
"""for i in range(2,10):
     b_col=worksheet.cell(row=i,column=2).value
     print(b_col)"""                                  #10 -1 = 9 sütun yazacak

880
1200
1500
8
560
780
920  #sonucuna ulaşmış oluyorum.

                        


#liste halinde bir sıralama istiyorum, iter_rows ve iter_cols kullanacağım.
"""rows = worksheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2)
print(rows)"""

#bu kısma kadar gelip çalıştırdığında <generator object Worksheet._cells_by_row at 0x00000210221FDEA0> şeklinde bir hata alıyorsun. bu hata bir hücre aralığını doğrudan yazdırmaya çalıştığınızda ortaya çıkar. Bu durumda, Python size bir jeneratör nesnesi döndürür ve doğrudan bu nesneyi yazdırmaya çalıştığınızda yukarıdaki mesajı alırsınız.
#bundan kurtulmak için devam ediyorum:

"""names = []      #names ve balance yerine istediğini yazabilirsin.
balance = []

for a,b in rows:
    names.append(a.value)
    balance.append(b.value)
    print(names)
    print(balance)"""

#printlerin konumundan kaynaklı şöyle bir cevap alıyorsun:
"""['Name']
['Balance']
['Name', 'John']
['Balance', 880]
['Name', 'John', 'Bob']
['Balance', 880, 1200]
['Name', 'John', 'Bob', 'Kane']
['Balance', 880, 1200, 1500]
['Name', 'John', 'Bob', 'Kane', 'Tim']
['Balance', 880, 1200, 1500, 1400]
['Name', 'John', 'Bob', 'Kane', 'Tim', 'Robin']
['Balance', 880, 1200, 1500, 1400, 560]
['Name', 'John', 'Bob', 'Kane', 'Tim', 'Robin', 'Michael']
['Balance', 880, 1200, 1500, 1400, 560, 780]"""

#bu çok mantıksız tabii. bu yüzden printleri döngü içerisine koymaman gerekiyor.

"""for a,b in rows:
    names.append(a.value)
    balance.append(b.value)

print(names)
print(balance)""" #sonucu:

['Name', 'John', 'Bob', 'Kane', 'Tim', 'Robin', 'Michael']
['Balance', 880, 1200, 1500, 1400, 560, 780]

#iter_cols için

"""columns = worksheet.iter_cols(min_row =1, max_row=5, min_col=1, max_col=2)
for col in columns:
    print(col)"""

"""(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.A2>, <Cell 'Sheet1'.A3>, <Cell 'Sheet1'.A4>, <Cell 'Sheet1'.A5>)
(<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.B4>, <Cell 'Sheet1'.B5>)"""

#şeklinde sonuç aldım.

"""rows = list(worksheet.rows)
print(rows)"""

#sonuç:

"""[(<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>), (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>), (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>), (<Cell 'Sheet1'.A4>, <Cell 'Sheet1'.B4>), (<Cell 'Sheet1'.A5>, <Cell 'Sheet1'.B5>), (<Cell 'Sheet1'.A6>, <Cell 'Sheet1'.B6>), (<Cell 'Sheet1'.A7>, <Cell 'Sheet1'.B7>), (<Cell 'Sheet1'.A8>, <Cell 'Sheet1'.B8>)]"""

#hücredeki değerin rengini, boyutunu, ve yazı tipini değiştirmek istiyorum:

"""from openpyxl.styles import Font,Color""" #başlangıç kısmına ekleme yapmayı unutma.

"""font_type = Font(name="istediğin yazı tipi",size= istediğin boyut ,color="rengin hex code u", italic= True,False, bold= True,False)"""

"""x=worksheet"""['değişmesini istediğin hücrenin konumu']
"""x.font = font_type"""

#bu şekilde değiştirebiliyorum.
#değişmesini istemediğin şeyi eklemene gerek yok.

#sütunu istediğim şekle getirmek istiyorum. yukarıdakinden farklı olarak eklemem gereken şey:

"""for i in range:""" #(kaçıncı hücreden, kaçıncı hücreye kadar kapsasın):
"""worksheet.cell(row=i, column= kaçıncı sütun istersen).font = font_type """ #(font type yerine istediğini yazabilirsin.)

#çalışan kod örneği:

"""font_type = (Font(name="Times New Roman", size=10, color="DB3B22",underline="single",bold=True ))"""

"""for i in range (2,10):
    worksheet.cell(row=i ,column=3).font = font_type"""

#strikethrough = üstünü çizmek
#underline komutunun single, double, singleAccounting, doubleAccounting gibi çeşitleri var.

# hücrenin arka planını renklendirmek ve desenle doldurmak istiyorum:
#başlangıca:
"""from openpyxl.styles import PatternFill""" #ekliyorum.

"""pattern_fill(istediğini yazabilirsin sadece tanımlaman gerek) = PatternFill(patternType ="istediğin şekil", fgcolor ="renk kodu")
worksheet["değişmesini istediğin hücrenin konumu"].fill = pattern_fill""" #tanımladığın şey bu anlama geliyor diyorsun aslında.

#hücreye çizgi eklemek istiyorum:
#öncelikle başına:
"""from openpyxl.styles import Border, Side"""
#ekliyorum.

"""kesikliçizgi(istediğini tanımlayabilirsin)= Side("dashed(istediğin çizgi tipi)",color = (renk kodu))"""

"""border(yine istediğini yazabilirsin)=Border(top=kesikliçizgi)
worksheet["istediğin hücre"].border = Border"""

#anlamazsan aşağıdaki kod çalışıyor:

"""kesikliçizgi = Side("dashed", color="FF0707")

border = Border(top=kesikliçizgi)

worksheet["B3"].border = border"""

#başka bir versiyonu:

"""kesikliçizgi = Side("dashed", color="FF0707")
yeşilçizgi = Side("double", color="10AF2A")
border = Border(top=kesikliçizgi, bottom=yeşilçizgi)
worksheet["A5"].border = border"""

#hücrelerin formatlarını yani hücreleri biçimlendir sekmesini nasıl değiştirebilirim?

"""worksheet["C4"].value= "11/11/20"
worksheet["D4"].value = "20"
worksheet["E4"].value = "Beginner" """

"""worksheet["E4"] = number_format = numbers.FORMAT_NUMBER
worksheet["C4"] = number_format = numbers.FORMAT_PERCENTAGE"""

#başlangıca eklemen gereken:
"""from openpyxl.styles import numbers"""
#hücrelere değer atadım 
#daha sonrasında formülü gördüğün gibi yazdım. sondaki formatlar dilediğin gibi değiştirebilirsin. zaten tıklayınca bir sürü çeşidi beliriyor.

#2 hücreden değer alıp çarpıp başka bir sütuna yazdırıyorum:
"""for i in range (2,10):
    balance = worksheet.cell(row=i, column=2).value
    interest = worksheet.cell(row=i, column=3).value
    final_balance = (balance*interest)+balance
    worksheet.cell(row=i, column=4).value = final_balance"""

#balance, interest ve final balance yerine istediğini yazabilirsin. balance ve interest sütunundan verileri alıp çarpıp balance ile toplayıp istediğim sütuna yazdırıyorum.












