from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

data = {
    "Joe": {
        "Matematik": 65,
        "Fizik": 72,
        "İngilizce": 45,
        "Beden": 100
    },
    "Dilek": {
        "Matematik": 42,
        "Fizik": 26,
        "İngilizce": 84,
        "Beden": 95
    },
    "Emir": {
        "Matematik": 52,
        "Fizik": 41,
        "İngilizce": 68,
        "Beden": 73
    },
    "Meric": {
        "Matematik": 89,
        "Fizik": 43,
        "İngilizce": 42,
        "Beden": 67
        }
}

workbook = Workbook()
worksheet = workbook.active
worksheet.title = "Grades"

headings = ['Name'] + list(data["Joe"].keys())
worksheet.append(headings)

for person in data:
    grades = list(data[person].values())
    worksheet.append([person] + grades)

for col in range(2, len(data["Joe"]) + 2):
    char = get_column_letter(col)
    worksheet[char + "6"] = f"=SUM({char + '2'}:{char + '5'})/{len(data)}"

for col in range(1, 6):
    worksheet[get_column_letter(col)+ '1'].font = Font(bold=True, color="0099CCFF")



workbook.save("C:\\Users\\kadir\\Desktop\\Python Excel Tutorial - Kendi Denemem.xlsx")
