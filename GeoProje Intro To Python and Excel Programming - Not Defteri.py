import openpyxl
from openpyxl import Workbook

#create a workbook object
workbook = Workbook()

#create an active workheet
worksheet = Workbook.active

#load existing spreadsheet
Workbook = openpyxl.load_workbook ("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")

# yukarıdaki kodu aynı şekilde çalıştırıp daha az kod yazmak amacıyla şu şekilde de yazabilirim:

import openpyxl #aslında burada bütün kütüphaneyi çekmiş oluyorum ama spesifik olarak bir komut çağırmak istediğimde from openpyxl import workbook gibi yazıp çağırabiliyorum.
                #bunu uyguladığımda workbook yazıldığında kodu direkt tanımlamış oluyorum. eğer öyle yazmazsam başına aşağıdaki gibi openpyxl eklemem gerekir ki tanısın.

#create a workbook object
workbook = openpyxl.Workbook()

#create an active workheet
worksheet = openpyxl.Workbook.active

#load existing spreadsheet
Workbook = openpyxl.load_workbook ("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")

#her seferinde tek tek openpyxl yazmak yerine daha pratik ve kısa gözükmesi açısından import openpyxl as "istediğin bir kısaltma ama genelde oe ya da op tercih ediliyor." şeklinde bir kısaltma atayabilirsin:

import openpyxl as kadiremir

#create a workbook object
workbook = kadiremir.Workbook()

#create an active workheet
worksheet = kadiremir.Workbook.active

#load existing spreadsheet
Workbook = openpyxl.load_workbook ("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")

#line 38' de gördüğün üzere openpyxl i import ettiğim için onu da tanımlıyor.

#print something from our spreadsheet
worksheet = workbook["Sheet"]
print(worksheet["A1"].value) 

#şimdi excel tablosundaki bilgileri örneğin john: blue şeklinde alt alta yazdıracağım.

import openpyxl

# Dosyayı aç
file_path = "GeoProje Intro To Python and Excel Programming.xlsx"
workbook = openpyxl.load_workbook("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")

# İlk sayfayı seç
worksheet = workbook.active

# Tüm hücrelerin değerlerini okuyup, "isim: renk" formatında yazdır
for row in worksheet.iter_rows(min_row=2, values_only=True):  # Başlık satırını atlamak için min_row=2
    if row[0] and row[1]:  # Eğer hücrede değer varsa
        print(f"{row[0]}: {row[1]}")
        
#John: Blue
#Erin: Red
#Sam: Pink
#Tina: Green            #şeklinde liste oluşturacaktır.
#Josh: Yellow
#Mary: Black
#Bob: White
#Lisa: Purple
#Steve: Gray


#alttaki kod dizisinde ise belirli hücereleri seçip print kısmında da görüldüğü üzere yan yana yazdırmasını istedim.

import openpyxl
from openpyxl import Workbook

Workbook = openpyxl.load_workbook ("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")
worksheet = Workbook.active

#set a variable
name = worksheet["A2"].value
color = worksheet["B2"].value

#print something from our speadsheet
print(f"{name}:{color}")


#bu sefer de sütundaki maddeleri aralarında boşluk olacak şekilde liste halinde yazdırmak istiyorum.

import openpyxl
from openpyxl import Workbook

Workbook = openpyxl.load_workbook("C:\\Users\\kadir\\Desktop\\GeoProje\\GeoProje Intro To Python and Excel Programming\\GeoProje Intro To Python and Excel Programming.xlsx")
worksheet = Workbook.active

column_a = worksheet["A"]
for cell in column_a:
    print(f"{cell}\n")          #\n yeni satır ekliyor. {cell.value} dersem hücredeki değerleri listeletmiş olurum.
                                #print(cell.value) dersem yine hücreler içerisindeki değerleri listeleyecek fakat değerler arasında boşluk bırakmayacaktır.


#yine aralarında boşluk kalmayacak şekilde listeliyorum ama farklı bir kod dizilimiyle:

range = worksheet["A2:A10"]

for cell in range:
    for x in cell:
        print(x.value)

#sütun değerlerini istediğin gibi değiştirebilirsin.


