"""import openpyxl
from openpyxl import Workbook


dosya = Workbook()
sayfa = dosya.active
sayfa["A1"].value = "python"
sayfa["A2"].value = "excel"

sayfa.title ="hatalar"

dosya.create_sheet("ikinci", 1)
dosya.create_sheet("ucuncu" ,2)

dosya.save("./yeni.xlsx")"""