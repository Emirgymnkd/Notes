"""import openpyxl
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill

dosya = openpyxl.load_workbook ("C:\\Users\\kadir\\Desktop\\data.xlsx")
sayfa = dosya ["notlar"]

sayfa.cell(row=1, column=4, value="Ortalama")
sayfa.cell(row=1, column=4).font = Font(name ="Calibri",sz=16)
dosya.save("C:\\Users\\kadir\\Desktop\\data.xlsx")

satir_sayisi = sayfa.max_row
for satir in range (2,satir_sayisi+1):
    vize = sayfa.cell(satir,2).value
    final = sayfa.cell(satir,3).value
    ortalama =(vize+final)/2
    sayfa.cell(row=satir, column=4, value=ortalama)
    sayfa.cell(satir,4).font = Font(name ="Calibri",sz=16, color="2C27CB")
    sayfa.cell(satir,4).fill = PatternFill("solid","27CB5B")
    if ortalama>70:
        sayfa.cell(satir,4).fill = PatternFill("solid","08D84A")
    else:
        sayfa.cell(satir,4).fill = PatternFill("solid", "E3331B")

dosya.save("C:\\Users\\kadir\\Desktop\\data.xlsx")"""